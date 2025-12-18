from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook

from .schemas import GOOD_STOCKS_COLUMNS, ordered_df


GOOD_SHEET = "GoodStocks"
RUNLOG_SHEET = "RunLog"


def _ensure_workbook(path: Path) -> None:
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(GOOD_SHEET)
        ws.title = GOOD_SHEET
        wb.create_sheet(RUNLOG_SHEET)
        wb.save(path)


def _to_float_or_none(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except Exception:
        pass
    try:
        return float(val)
    except Exception:
        return None


def write_good_stocks(excel_path: str | Path, df: pd.DataFrame) -> None:
    path = Path(excel_path)
    _ensure_workbook(path)

    # Normalize/validate columns so Excel output is deterministic.
    df = df.copy() if df is not None else pd.DataFrame()
    if not df.empty:
        df = ordered_df(df, [c for c in GOOD_STOCKS_COLUMNS if c != "Timestamp"])

    wb = load_workbook(path)
    if GOOD_SHEET in wb.sheetnames:
        ws = wb[GOOD_SHEET]
        wb.remove(ws)
        wb.create_sheet(GOOD_SHEET)
    else:
        wb.create_sheet(GOOD_SHEET)
    ws = wb[GOOD_SHEET]

    # Write header
    timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    ws.append(GOOD_STOCKS_COLUMNS)

    # Write rows
    for _, row in df.iterrows():
        ws.append([
            timestamp,
            row.get("Ticker"),
            row.get("Strategy"),
            _to_float_or_none(row.get("Score", 0.0)),
            _to_float_or_none(row.get("Close")),
            _to_float_or_none(row.get("RSI14")),
            _to_float_or_none(row.get("SMA50")),
            _to_float_or_none(row.get("SMA200")),
            _to_float_or_none(row.get("Dist_52wHigh")),
            _to_float_or_none(row.get("Vol5x20")),
        ])

    # RunLog
    if RUNLOG_SHEET not in wb.sheetnames:
        wb.create_sheet(RUNLOG_SHEET)
    log = wb[RUNLOG_SHEET]
    if log.max_row == 1 and log.max_column == 1 and log.cell(1, 1).value is None:
        log.append(["Timestamp", "Candidates", "Notes"])
    log.append([timestamp, int(len(df)), "Auto scan update"])

    wb.save(path)
